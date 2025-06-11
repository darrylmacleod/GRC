import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import json
import logging
from typing import Dict, List, Optional, Any
from difflib import get_close_matches
from enum import Enum
import openpyxl
import threading

try:
    from rapidfuzz import process as rapidfuzz_process
    FUZZY_LIB = 'rapidfuzz'
except ImportError:
    FUZZY_LIB = 'difflib'

logging.basicConfig(
    filename='dataclassifier.log',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s'
)

class ClassificationLevel(Enum):
    PUBLIC = "Public"
    INTERNAL = "Internal"
    CONFIDENTIAL = "Confidential"
    RESTRICTED = "Restricted"

    @classmethod
    def list(cls) -> List[str]:
        return [e.value for e in cls]

class ClassificationMap:
    """
    Handles classification data and logic.
    """
    def __init__(self):
        self.classification_map: Dict[str, str] = {
            "Company website content": ClassificationLevel.PUBLIC.value,
            "Marketing materials": ClassificationLevel.PUBLIC.value,
            "Press releases": ClassificationLevel.PUBLIC.value,
            "Published research papers": ClassificationLevel.PUBLIC.value,
            "Public regulatory filings": ClassificationLevel.PUBLIC.value,
            "Internal memos and communications": ClassificationLevel.INTERNAL.value,
            "Organization charts": ClassificationLevel.INTERNAL.value,
            "Training materials": ClassificationLevel.INTERNAL.value,
            "Project plans and status reports": ClassificationLevel.INTERNAL.value,
            "Meeting agendas and minutes (non-sensitive)": ClassificationLevel.INTERNAL.value,
            "Employee ID numbers": ClassificationLevel.CONFIDENTIAL.value,
            "Business strategies": ClassificationLevel.CONFIDENTIAL.value,
            "Contract details": ClassificationLevel.CONFIDENTIAL.value,
            "Internal financial statements": ClassificationLevel.CONFIDENTIAL.value,
            "Customer contact information": ClassificationLevel.CONFIDENTIAL.value,
            "Source code": ClassificationLevel.CONFIDENTIAL.value,
            "Non-public pricing or product roadmaps": ClassificationLevel.CONFIDENTIAL.value,
            "Vendor agreements": ClassificationLevel.CONFIDENTIAL.value,
            "Intellectual property documentation": ClassificationLevel.CONFIDENTIAL.value,
            "Full names, addresses, phone numbers": ClassificationLevel.RESTRICTED.value,
            "Social Insurance Numbers (SIN)/Social Security Numbers (SSN)": ClassificationLevel.RESTRICTED.value,
            "Driver’s license numbers": ClassificationLevel.RESTRICTED.value,
            "Dates of birth": ClassificationLevel.RESTRICTED.value,
            "Passport numbers": ClassificationLevel.RESTRICTED.value,
            "Medical records": ClassificationLevel.RESTRICTED.value,
            "Health insurance data": ClassificationLevel.RESTRICTED.value,
            "Lab test results": ClassificationLevel.RESTRICTED.value,
            "Appointment histories": ClassificationLevel.RESTRICTED.value,
            "Credit card numbers": ClassificationLevel.RESTRICTED.value,
            "CVV codes": ClassificationLevel.RESTRICTED.value,
            "Cardholder names and billing addresses": ClassificationLevel.RESTRICTED.value,
            "Bank account numbers": ClassificationLevel.RESTRICTED.value,
            "Routing numbers": ClassificationLevel.RESTRICTED.value,
            "Tax returns": ClassificationLevel.RESTRICTED.value,
            "Passwords": ClassificationLevel.RESTRICTED.value,
            "API keys": ClassificationLevel.RESTRICTED.value,
            "Encryption keys": ClassificationLevel.RESTRICTED.value,
            "Biometric data": ClassificationLevel.RESTRICTED.value,
            "Litigation documents": ClassificationLevel.RESTRICTED.value,
            "Legal holds": ClassificationLevel.RESTRICTED.value,
            "Regulatory investigation materials": ClassificationLevel.RESTRICTED.value,
            "Board meeting minutes": ClassificationLevel.RESTRICTED.value,
            "Merger/acquisition plans": ClassificationLevel.RESTRICTED.value,
            "Due diligence documents": ClassificationLevel.RESTRICTED.value,
        }
        self.classification_levels: set[str] = set(ClassificationLevel.list())

    def add_classification(self, data_type: str, level: str) -> bool:
        """Add a new data type and classification level. Return True if added, False if invalid or duplicate."""
        if not data_type or not level:
            logging.warning(f"[add_classification] Invalid: {data_type}, {level}")
            return False
        if data_type in self.classification_map:
            logging.warning(f"[add_classification] Duplicate: {data_type}")
            return False
        self.classification_map[data_type] = level
        self.classification_levels.add(level)
        logging.info(f"[add_classification] Added: {data_type} as {level}")
        return True

    def remove_classification(self, data_type: str) -> bool:
        """Remove a data type from the classification map. Return True if removed, False if not found."""
        if data_type in self.classification_map:
            del self.classification_map[data_type]
            logging.info(f"[remove_classification] Removed: {data_type}")
            return True
        logging.warning(f"[remove_classification] Not found: {data_type}")
        return False

    def import_csv(self, file_path: str) -> None:
        try:
            with open(file_path, mode='r', encoding='utf-8') as file:
                reader = csv.reader(file)
                next(reader, None)
                for row in reader:
                    if len(row) == 2:
                        self.add_classification(row[0], row[1])
                    else:
                        logging.warning(f"[import_csv] Skipped row: {row}")
        except Exception as e:
            logging.exception(f"[import_csv] Error: {e}")
            raise

    def export_csv(self, file_path: str) -> None:
        try:
            with open(file_path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(["Data Type", "Classification"])
                for data_type, level in self.classification_map.items():
                    writer.writerow([data_type, level])
        except Exception as e:
            logging.exception(f"[export_csv] Error: {e}")
            raise

    def import_excel(self, file_path: str) -> None:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    self.add_classification(str(row[0]), str(row[1]))
                else:
                    logging.warning(f"[import_excel] Skipped row {i}: {row}")
        except Exception as e:
            logging.exception(f"[import_excel] Error: {e}")
            raise

    def export_excel(self, file_path: str) -> None:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Data Type", "Classification"])
            for data_type, level in self.classification_map.items():
                ws.append([data_type, level])
            wb.save(file_path)
        except Exception as e:
            logging.exception(f"[export_excel] Error: {e}")
            raise

    def import_json(self, file_path: str) -> None:
        try:
            with open(file_path, mode='r', encoding='utf-8') as file:
                data = json.load(file)
                for obj in data:
                    if isinstance(obj, dict) and "Data Type" in obj and "Classification" in obj:
                        self.add_classification(obj["Data Type"], obj["Classification"])
                    else:
                        logging.warning(f"[import_json] Skipped entry: {obj}")
        except Exception as e:
            logging.exception(f"[import_json] Error: {e}")
            raise

    def export_json(self, file_path: str) -> None:
        try:
            data = [{"Data Type": dt, "Classification": cl} for dt, cl in self.classification_map.items()]
            with open(file_path, mode='w', encoding='utf-8') as file:
                json.dump(data, file, indent=2)
        except Exception as e:
            logging.exception(f"[export_json] Error: {e}")
            raise

    def get_classification(self, data_type: str, fuzzy_threshold: int = 60) -> str:
        """Get classification for a data type, or fuzzy match if not found."""
        result = self.classification_map.get(data_type)
        if result:
            logging.info(f"[get_classification] Exact: '{data_type}' as '{result}'")
            return result
        keys = list(self.classification_map.keys())
        if FUZZY_LIB == 'rapidfuzz':
            matches = rapidfuzz_process.extract(data_type, keys, limit=1, score_cutoff=fuzzy_threshold)
            if matches:
                logging.info(f"[get_classification] Fuzzy: '{data_type}' as '{matches[0][0]}'")
                return self.classification_map[matches[0][0]]
        else:
            matches = get_close_matches(data_type, keys, n=1, cutoff=fuzzy_threshold / 100.0)
            if matches:
                logging.info(f"[get_classification] Fuzzy: '{data_type}' as '{matches[0]}'")
                return self.classification_map[matches[0]]
        logging.warning(f"[get_classification] No match: '{data_type}'")
        return "Unknown"

    def get_all_types(self) -> List[str]:
        return list(self.classification_map.keys())

    def get_levels(self) -> List[str]:
        return list(self.classification_levels)

class ThemeManager:
    """Handles theme application for the GUI."""
    LIGHT = {
        "bg": "#FFFFFF",
        "fg": "#333",
        "Public": "green",
        "Internal": "#2057B3",
        "Confidential": "#D37D00",
        "Restricted": "#C00"
    }
    DARK = {
        "bg": "#232323",
        "fg": "#F0F0F0",
        "Public": "#55D46A",
        "Internal": "#6AC6FF",
        "Confidential": "#FFD16A",
        "Restricted": "#FF7575"
    }

    def __init__(self):
        self.theme = "light"

    def toggle(self) -> None:
        self.theme = "dark" if self.theme == "light" else "light"

    def get_color(self, level: str) -> str:
        palette = self.DARK if self.theme == "dark" else self.LIGHT
        return palette.get(level, palette["fg"])

    def get_bg(self) -> str:
        return self.DARK["bg"] if self.theme == "dark" else self.LIGHT["bg"]

    def is_dark(self) -> bool:
        return self.theme == "dark"

class DataClassifierApp:
    """
    Tkinter GUI for the data classification tool.
    """
    def __init__(self, root: tk.Tk, cmap: ClassificationMap):
        self.root: tk.Tk = root
        self.cmap: ClassificationMap = cmap
        self.theme_mgr: ThemeManager = ThemeManager()
        self.fuzzy_threshold: int = 60
        self.root.title("Data Classification Tool")
        style = ttk.Style()
        style.theme_use('clam')
        self.setup_style(style)
        self.setup_main_layout()
        self.setup_menu()
        self.setup_search_section()
        self.setup_add_section()
        self.setup_export_import_section()
        self.setup_status_bar()
        self.apply_theme()
        self.setup_shortcuts()

    def setup_style(self, style: ttk.Style) -> None:
        style.configure("TLabel", font=("Segoe UI", 11))
        style.configure("TButton", font=("Segoe UI", 10), padding=6)
        style.configure("TEntry", font=("Segoe UI", 10))
        style.configure("TCombobox", font=("Segoe UI", 10))
        style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))

    def setup_main_layout(self) -> None:
        self.main = ttk.Frame(self.root, padding="18 16 18 16")
        self.main.grid(row=0, column=0, sticky="NSEW")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main.columnconfigure(0, weight=1)

    def setup_menu(self) -> None:
        menubar = tk.Menu(self.root)
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="How to Use", command=self.show_help)
        helpmenu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=helpmenu)
        thememenu = tk.Menu(menubar, tearoff=0)
        thememenu.add_command(label="Toggle Light/Dark", command=self.toggle_theme)
        menubar.add_cascade(label="Theme", menu=thememenu)
        self.root.config(menu=menubar)

    def setup_search_section(self) -> None:
        row = 0
        ttk.Label(self.main, text="Search Data Type:").grid(row=row, column=0, sticky="W")
        row += 1
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_dropdown)
        self.search_entry = ttk.Entry(self.main, textvariable=self.search_var, width=50)
        self.search_entry.grid(row=row, column=0, sticky="EW", pady=(0, 8))
        self.add_tooltip(self.search_entry, "Type to search for a data type.")
        row += 1
        self.data_type_var = tk.StringVar()
        self.dropdown = ttk.Combobox(self.main, textvariable=self.data_type_var, width=60)
        self.dropdown['values'] = self.cmap.get_all_types()
        self.dropdown.grid(row=row, column=0, sticky="EW", pady=(0, 8))
        self.dropdown.bind("<<ComboboxSelected>>", self.classify_data)
        self.add_tooltip(self.dropdown, "Select a data type to see its classification.")
        row += 1
        self.result_label = ttk.Label(self.main, text="", font=("Arial", 12, "bold"), anchor="center")
        self.result_label.grid(row=row, column=0, pady=(0, 12), sticky="EW")
        row += 1
        ttk.Separator(self.main, orient="horizontal").grid(row=row, column=0, sticky="EW", pady=(0,12))
        self.search_section_row = row + 1

    def setup_add_section(self) -> None:
        row = self.search_section_row
        add_frame = ttk.LabelFrame(self.main, text="Add/Remove Custom Data Type", padding="10 8 10 8")
        add_frame.grid(row=row, column=0, sticky="EW", padx=0, pady=(0,12))
        add_frame.columnconfigure(0, weight=1)
        self.new_data_var = tk.StringVar()
        self.new_classification_var = tk.StringVar()
        entry = ttk.Entry(add_frame, textvariable=self.new_data_var, width=40)
        entry.grid(row=0, column=0, sticky="EW", pady=(0, 8))
        self.add_tooltip(entry, "Enter a new data type.")
        self.new_class_dropdown = ttk.Combobox(
            add_frame, textvariable=self.new_classification_var,
            values=self.cmap.get_levels(), width=37
        )
        self.new_class_dropdown.grid(row=1, column=0, sticky="EW", pady=(0, 8))
        self.add_tooltip(self.new_class_dropdown, "Select a classification level.")
        btn = ttk.Button(add_frame, text="Add Data Type", command=self.add_custom_type)
        btn.grid(row=2, column=0, sticky="EW")
        self.add_tooltip(btn, "Add the custom data type to the list.")
        # --- Remove data type button ---
        remove_btn = ttk.Button(add_frame, text="Remove Selected Data Type", command=self.remove_selected_type)
        remove_btn.grid(row=3, column=0, sticky="EW", pady=(8, 0))
        self.add_tooltip(remove_btn, "Remove the selected data type from the list.")
        self.add_section_row = row + 1

    def setup_export_import_section(self) -> None:
        row = self.add_section_row
        export_frame = ttk.Frame(self.main)
        export_frame.grid(row=row, column=0, sticky="EW")
        ttk.Button(export_frame, text="Export Classifications to CSV", command=self.export_csv).grid(row=0, column=0, sticky="EW", padx=(0,8))
        ttk.Button(export_frame, text="Import Classifications from CSV", command=self.import_csv).grid(row=0, column=1, sticky="EW")
        ttk.Button(export_frame, text="Export to Excel", command=self.export_excel).grid(row=1, column=0, sticky="EW", padx=(0,8), pady=(5,0))
        ttk.Button(export_frame, text="Import from Excel", command=self.import_excel).grid(row=1, column=1, sticky="EW", pady=(5,0))
        ttk.Button(export_frame, text="Export as JSON", command=self.export_json).grid(row=2, column=0, sticky="EW", padx=(0,8), pady=(5,0))
        ttk.Button(export_frame, text="Import from JSON", command=self.import_json).grid(row=2, column=1, sticky="EW", pady=(5,0))
        export_frame.columnconfigure(0, weight=1)
        export_frame.columnconfigure(1, weight=1)
        self.export_section_row = row + 1

    def setup_status_bar(self) -> None:
        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(self.main, textvariable=self.status_var, anchor="w", relief="sunken")
        self.status_bar.grid(row=self.export_section_row, column=0, sticky="EW")
        self.set_status("Ready.")

    def setup_shortcuts(self) -> None:
        self.root.bind('<Control-e>', lambda e: self.export_csv())
        self.root.bind('<Control-i>', lambda e: self.import_csv())
        self.root.bind('<Control-x>', lambda e: self.export_excel())
        self.root.bind('<Control-j>', lambda e: self.export_json())
        self.root.bind('<Control-t>', lambda e: self.toggle_theme())
        self.root.bind('<Control-f>', lambda e: self.search_entry.focus_set())

    def update_dropdown(self, *args) -> None:
        query = self.search_var.get().lower()
        all_types = self.cmap.get_all_types()
        if not query:
            filtered = all_types
        else:
            matches = [k for k in all_types if query in k.lower()]
            if FUZZY_LIB == 'rapidfuzz':
                fuzzy_matches = [m[0] for m in rapidfuzz_process.extract(query, all_types, limit=10, score_cutoff=self.fuzzy_threshold)]
            else:
                fuzzy_matches = get_close_matches(query, all_types, n=10, cutoff=self.fuzzy_threshold / 100.0)
            filtered = list(dict.fromkeys(matches + fuzzy_matches))
        self.dropdown['values'] = filtered

    def classify_data(self, event=None) -> None:
        selected_data = self.data_type_var.get()
        classification = self.cmap.get_classification(selected_data, fuzzy_threshold=self.fuzzy_threshold)
        if classification == "Unknown":
            self.result_label.configure(text="Classification: Unknown", foreground="#555")
            self.set_status(f"Unknown classification for '{selected_data}'", error=True)
        else:
            color = self.theme_mgr.get_color(classification)
            self.result_label.configure(
                text=f"Classification: {classification}",
                foreground=color
            )
            self.set_status(f"Classified '{selected_data}' as '{classification}'.")

    def add_custom_type(self) -> None:
        new_data = self.new_data_var.get().strip()
        new_class = self.new_classification_var.get().strip()
        if not new_data or not new_class:
            self.set_status("Please enter a valid data type and classification level.", error=True)
            return
        if self.cmap.add_classification(new_data, new_class):
            self.dropdown['values'] = self.cmap.get_all_types()
            self.new_class_dropdown['values'] = self.cmap.get_levels()
            self.result_label.configure(
                text=f"Added: {new_data} as {new_class}",
                foreground=self.theme_mgr.get_color(new_class)
            )
            self.set_status(f"Added: '{new_data}' as '{new_class}'.")
            self.new_data_var.set("")
            self.new_classification_var.set("")
        else:
            self.set_status("Failed to add the data type (may be duplicate or invalid).", error=True)

    def remove_selected_type(self) -> None:
        selected_data = self.data_type_var.get().strip()
        if not selected_data:
            self.set_status("Please select a data type to remove.", error=True)
            return
        confirm = messagebox.askyesno(
            "Confirm Removal", f"Are you sure you want to remove '{selected_data}'?"
        )
        if not confirm:
            self.set_status("Removal cancelled.")
            return
        if self.cmap.remove_classification(selected_data):
            self.dropdown['values'] = self.cmap.get_all_types()
            self.new_class_dropdown['values'] = self.cmap.get_levels()
            self.result_label.configure(text=f"Removed: {selected_data}", foreground="red")
            self.set_status(f"Removed '{selected_data}' from the list.")
            self.data_type_var.set("")
        else:
            self.set_status(f"Data type '{selected_data}' not found.", error=True)

    def _run_in_thread(self, func, *args):
        def task():
            try:
                func(*args)
                self.set_status("Done.")
            except Exception as e:
                self.set_status(f"Operation failed: {e}", error=True)
        threading.Thread(target=task, daemon=True).start()

    def export_csv(self) -> None:
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if file_path:
            self.set_status("Exporting to CSV...")
            self._run_in_thread(self.cmap.export_csv, file_path)

    def import_csv(self) -> None:
        file_path = filedialog.askopenfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if file_path:
            self.set_status("Importing from CSV...")
            def task():
                self.cmap.import_csv(file_path)
                self.dropdown['values'] = self.cmap.get_all_types()
                self.new_class_dropdown['values'] = self.cmap.get_levels()
            self._run_in_thread(task)

    def export_excel(self) -> None:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.set_status("Exporting to Excel...")
            self._run_in_thread(self.cmap.export_excel, file_path)

    def import_excel(self) -> None:
        file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.set_status("Importing from Excel...")
            def task():
                self.cmap.import_excel(file_path)
                self.dropdown['values'] = self.cmap.get_all_types()
                self.new_class_dropdown['values'] = self.cmap.get_levels()
            self._run_in_thread(task)

    def export_json(self) -> None:
        file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if file_path:
            self.set_status("Exporting to JSON...")
            self._run_in_thread(self.cmap.export_json, file_path)

    def import_json(self) -> None:
        file_path = filedialog.askopenfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if file_path:
            self.set_status("Importing from JSON...")
            def task():
                self.cmap.import_json(file_path)
                self.dropdown['values'] = self.cmap.get_all_types()
                self.new_class_dropdown['values'] = self.cmap.get_levels()
            self._run_in_thread(task)

    def show_help(self) -> None:
        help_text = (
            "How to Use the Data Classification Tool:\n"
            "- Search for a data type using the search box.\n"
            "- Select a data type to see its classification.\n"
            "- Add or remove a custom data type below.\n"
            "- Import/export data using the buttons.\n"
            "- Use the Theme menu or Ctrl+T to toggle light/dark mode.\n"
            "- Keyboard shortcuts:\n"
            "  Ctrl+E: Export CSV\n"
            "  Ctrl+I: Import CSV\n"
            "  Ctrl+X: Export Excel\n"
            "  Ctrl+J: Export JSON\n"
            "  Ctrl+F: Focus search\n"
        )
        messagebox.showinfo("Help", help_text)

    def show_about(self) -> None:
        messagebox.showinfo(
            "About",
            "Data Classification Tool\n"
            "Author: darrylmacleod\n"
            "Version: 2.0 (Rewritten)\n"
            "GitHub: https://github.com/darrylmacleod/GRC"
        )

    def toggle_theme(self) -> None:
        self.theme_mgr.toggle()
        self.apply_theme()

    def apply_theme(self) -> None:
        bg = self.theme_mgr.get_bg()
        self.root.configure(bg=bg)
        self.main.configure(style="Dark.TFrame" if self.theme_mgr.is_dark() else "TFrame")
        for widget in self.root.winfo_children():
            try:
                widget.configure(bg=bg)
            except Exception:
                pass

    def set_status(self, message: str, error: bool = False) -> None:
        self.status_var.set(message)
        if error:
            self.status_bar.configure(foreground="red")
            logging.error(message)
            messagebox.showerror("Error", message)
        else:
            self.status_bar.configure(foreground="green")
            logging.info(message)

    def add_tooltip(self, widget, text: str) -> None:
        def on_enter(event):
            x, y, cx, cy = widget.bbox("insert")
            x += widget.winfo_rootx() + 25
            y += widget.winfo_rooty() + 20
            self.tipwindow = tw = tk.Toplevel(widget)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x}+{y}")
            label = tk.Label(
                tw, text=text, background="#ffffe0",
                relief='solid', borderwidth=1, font=("tahoma", "8", "normal")
            )
            label.pack(ipadx=1)
        def on_leave(event):
            if hasattr(self, 'tipwindow') and self.tipwindow:
                self.tipwindow.destroy()
                self.tipwindow = None
        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

if __name__ == "__main__":
    root = tk.Tk()
    cmap = ClassificationMap()
    app = DataClassifierApp(root, cmap)
    root.mainloop()
